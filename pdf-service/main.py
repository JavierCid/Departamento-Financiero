from fastapi import FastAPI, UploadFile, File, HTTPException, Response, Request

from fastapi.middleware.cors import CORSMiddleware
from typing import List
from urllib.parse import quote
import io
import json
import pandas as pd
import pdfplumber
from extractor import extract_from_pages
from datetime import datetime
from bankflow_rules import process_bankflow  # ← Usamos el pipeline completo
import uvicorn

app = FastAPI()

# CORS para que Blazor pueda llamar al servicio en local
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],      # en local nos da igual el origen
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



# 🔧 Aceptar cualquier origen en local (localhost / 127.0.0.1)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:7252",
        "https://localhost:7252",
        "http://127.0.0.1:7252",
        "https://127.0.0.1:7252",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition", "X-Preview"],
)


@app.options("/{full_path:path}")
async def cors_preflight(full_path: str):
    return Response(status_code=204)


# =========================


def _clip(text: str | None, maxlen: int = 27) -> str:
    s = "" if text is None else str(text)
    return s if len(s) <= maxlen else s[: maxlen - 3] + "..."


def _short_name(filename: str, maxlen: int = 27) -> str:
    if not filename:
        return ""
    if len(filename) <= maxlen:
        return filename
    parts = filename.rsplit(".", 1)
    if len(parts) == 2:
        stem, ext = parts[0], "." + parts[1]
    else:
        stem, ext = filename, ""
    keep = maxlen - len(ext) - 3
    if keep < 1:
        return "..." + ext
    return stem[:keep] + "..." + ext


def _fmt_eur(v):
    try:
        if v is None or v == "" or pd.isna(v):
            return "—"
        n = float(v)
        return f"{n:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v) if v is not None else "—"

# ==== Helpers BankFlow Pro ====


def _read_tabular(upload: UploadFile) -> pd.DataFrame:
    """
    Lee CSV o Excel y devuelve un DataFrame (todo en str).
    CSV: prueba separador ';' y luego ','.
    Excel: detecta la fila de encabezados por contenido,
           para saltar metadatos (logo, titular, cuenta, etc.).
    """
    name = (upload.filename or "").lower()
    raw = upload.file.read()
    bio = io.BytesIO(raw)

    if name.endswith(".csv"):
        try:
            df = pd.read_csv(bio, sep=";", dtype=str, encoding="utf-8", engine="python")
        except Exception:
            bio.seek(0)
            df = pd.read_csv(bio, sep=",", dtype=str, encoding="utf-8", engine="python")
        return df.fillna("")
    else:
        # Excel con posibles filas de metadatos arriba
        try:
            xls = pd.ExcelFile(bio, engine="openpyxl")
        except Exception as e:
            raise RuntimeError(f"Lectura Excel falló: {type(e).__name__}: {e}")

        import unicodedata

        def _norm_cell(x: str) -> str:
            s = str(x or "").strip().lower()
            s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
            return s.replace("\n", " ").replace("\r", " ")

        def _find_header_row(df_nohdr: pd.DataFrame) -> int | None:
            """
            Busca la fila de cabecera detectando patrones:
            1. Extracto: (fecha + concepto + importe)
            2. Remesa: (importe + (beneficiario/proveedor/nombre o concepto))
            """
            MAX_SCAN = min(30, len(df_nohdr))
            for i in range(MAX_SCAN):
                row_vals = [_norm_cell(v) for v in df_nohdr.iloc[i].tolist()]
                if not any(row_vals):
                    continue
                
                # --- Detectar palabras clave ---
                has_fecha = any("fecha" in v for v in row_vals)
                has_concepto = any("concepto" in v for v in row_vals)
                has_importe = any("importe" in v for v in row_vals)
                
                # Palabras clave de Remesa (Beneficiario/Proveedor)
                has_beneficiario = any(
                    any(k in v for k in ["beneficiario", "proveedor", "nombre", "destinatario"]) 
                    for v in row_vals
                )
                
                # --- Comprobar patrones ---
                
                # Patrón 1: Extracto bancario
                if has_fecha and has_concepto and has_importe:
                    return i
                    
                # Patrón 2: Detalle de remesa
                if has_importe and (has_beneficiario or has_concepto):
                    return i
                    
            return None # No se encontró ninguna cabecera conocida

        # Recorre hojas y detecta cabecera por contenido
        for sheet in xls.sheet_names:
            try:
                tmp = xls.parse(sheet_name=sheet, header=None, dtype=str)
            except Exception:
                continue

            hdr_row = _find_header_row(tmp)
            if hdr_row is not None:
                try:
                    # Lee el archivo usando la fila de cabecera detectada
                    df = xls.parse(sheet_name=sheet, header=hdr_row, dtype=str)
                    if df.shape[1] >= 2: # Solo necesita 2+ columnas
                        return df.fillna("") # Devuelvefillna("") aquí
                except Exception:
                    continue

        # Fallback: primera hoja tal cual (probablemente fallará pero es el último recurso)
        try:
            return xls.parse(xls.sheet_names[0], dtype=str).fillna("")
        except Exception as e:
            raise RuntimeError(f"Excel parse sin cabecera también falló: {type(e).__name__}: {e}")


def _norm_colnames(df: pd.DataFrame) -> pd.DataFrame:
    import unicodedata

    def norm(s):
        s = str(s or "").strip().lower()
        s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
        s = s.replace("\n", " ").replace("\r", " ")
        return s
    df = df.copy()
    df.columns = [norm(c) for c in df.columns]
    return df


def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in df.columns:
        for cand in candidates:
            if cand in c:
                return c
    return None


def _to_date_ddmmyyyy(txt: str) -> str:
    if not txt or str(txt).strip() == "":
        return ""
    t = str(txt).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(t, fmt).strftime("%d/%m/%Y")
        except Exception:
            continue
    try:
        d = pd.to_datetime(t, dayfirst=True, errors="coerce")
        if pd.notna(d):
            return d.strftime("%d/%m/%Y")
    except Exception:
        pass
    return t


def _to_float_eu(txt: str) -> float | None:
    if txt is None:
        return None
    s = str(txt).strip().replace("€", "").replace("EUR", "").replace(" ", "")
    # 1.234,56 -> 1234.56
    if "." in s and "," in s and s.rfind(",") > s.rfind("."):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        try:
            return float(s.replace(".", "").replace(",", "."))
        except Exception:
            return None


def _fmt_eu(v: float | None) -> str:
    if v is None or pd.isna(v):
        return ""
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# =========================
# Parser
# =========================
def parse_pdf_to_df(pdf_bytes: bytes, nombre_archivo: str) -> pd.DataFrame:
    """
    Usa el extractor estable (Neto + IVA + IRPF = Importe Bruto, tolerancia ±0,05),
    corrige el patrón “21,00 % I.V.A. s/…”, y limpia incoherencias.
    """
    # Extraer textos de todas las páginas
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        pages_texts = []
        for p in pdf.pages:
            t = p.extract_text() or ""
            if t.strip():
                pages_texts.append(t)

    # Pasar por el extractor
    fields = extract_from_pages(pages_texts, nombre_archivo)

    # Normalizar nombres
    row = {
        "Proveedor": fields.get("Proveedor"),
        "Fecha": fields.get("Fecha"),
        "Invoice": fields.get("Invoice") or fields.get("Factura") or fields.get("Nº factura"),
        "Concepto": fields.get("Concepto"),
        "Neto": fields.get("Neto"),
        "IVA": fields.get("IVA"),
        "IRPF": fields.get("IRPF"),
        "Importe Bruto": fields.get("Importe bruto") or fields.get("Total Bruto") or fields.get("Bruto"),
    }

    cols = ["Proveedor", "Fecha", "Invoice", "Concepto", "Neto", "IVA", "IRPF", "Importe Bruto"]
    return pd.DataFrame([row], columns=cols)

# =========================
# Endpoint principal
# =========================
@app.post("/api/pdf2excel")
async def pdf2excel(file: List[UploadFile] = File(...)):
    """
    Acepta uno o varios PDFs y devuelve un Excel + cabecera 'X-Preview'.
    """
    if not file:
        raise HTTPException(status_code=400, detail="Sube al menos un PDF")

    dfs: list[pd.DataFrame] = []
    for f in file:
        if not f.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail=f"'{f.filename}' no es un PDF")

        contenido = await f.read()
        if not contenido:
            raise HTTPException(status_code=400, detail=f"'{f.filename}' está vacío")

        try:
            df = parse_pdf_to_df(contenido, f.filename)
            # Añadimos columna Archivo (nombre completo) para Excel;
            # y versión recortada para vista previa
            df["Archivo"] = f.filename
            df["ArchivoPreview"] = _short_name(f.filename, 27)
            dfs.append(df)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error leyendo '{f.filename}': {e}")

    df_total = pd.concat(dfs, ignore_index=True) if len(dfs) > 1 else dfs[0]

    # ====== Generar Excel ======
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        # Orden de columnas para Excel (incluimos Archivo completo)
        cols = [
            "Archivo",
            "Proveedor",
            "Fecha",
            "Invoice",
            "Concepto",
            "Neto",
            "IVA",
            "IRPF",
            "Importe Bruto",
        ]
        df_excel = df_total.reindex(columns=cols)
        df_excel.to_excel(w, index=False, sheet_name="Facturas")

        # Formato bonito
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = w.book
        ws = wb["Facturas"]

        # Formato numérico europeo (€ con punto de miles y coma decimal)
        euro_fmt = '#,##0.00 [$€-40C]'
        for col in ["F", "G", "H", "I"]:  # Neto, IVA, IRPF, Importe Bruto
            for cell in ws[col]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = euro_fmt

        # Quitar cuadrícula
        ws.sheet_view.showGridLines = False

        # Estilos encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Ajuste de ancho y bordes/alineación general
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
                # Bordes + alineación por tipo
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions[col_letter].width = max_len + 4

    xlsx_bytes = out.getvalue()

    # ====== Vista previa (máx. 50 filas) ======
    preview_rows = []
    for _, r in df_total.head(50).iterrows():
        preview_rows.append({
            "Archivo": _clip(r.get("ArchivoPreview"), 27),     # <= 27
            "OCR": "—",
            "Proveedor": _clip(r.get("Proveedor"), 27),        # <= 27
            "Fecha": r.get("Fecha"),
            "Invoice": r.get("Invoice"),
            "Concepto": r.get("Concepto"),
            "Total Neto": _fmt_eur(r.get("Neto")),
            "IVA €": _fmt_eur(r.get("IVA")),
            "IRPF": _fmt_eur(r.get("IRPF")),
            "Total Bruto": _fmt_eur(r.get("Importe Bruto")),
        })

    preview = {"Filas": int(len(df_total)), "Muestra": preview_rows}

    # ====== Nombre de salida ======
    base = (file[0].filename if file else "archivo.pdf").rsplit(".", 1)[0]
    out_name = f"Desglose_{base}.xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    headers = {
        "Content-Disposition": f'attachment; filename="{out_name}"; filename*=UTF-8\'\'{quote(out_name)}',
        # ensure_ascii=True evita problemas de codificación en cabeceras
        "X-Preview": json.dumps(preview, ensure_ascii=True),
    }

    return Response(content=xlsx_bytes, media_type=content_type, headers=headers)
# =========================
# Endpoint BankFlow Pro
# =========================
# (…tu endpoint /api/bankflowpro tal y como lo tienes…)

# =========================
# Endpoint: Contraste facturas (PDFs) vs Pendientes (Excel)
# =========================
from fastapi import UploadFile
import re

def _norm_invoice_code(s: str | None) -> str:
    """
    Normaliza códigos de factura sin destruir su estructura.
    - Conserva dígitos largos (3020014885)
    - Une letras+números (MA1391)
    - Quita solo ruido textual tipo 'Factura', 'Nº', 'Invoice', etc.
    """
    if not s:
        return ""
    import re
    t = str(s).upper().strip()

    # 1️⃣ Eliminar prefijos inútiles
    t = re.sub(r"\b(FACTURA|FAC|N[ºO]?|INVOICE|NO|NUMERO|NÚMERO|DOC|DOCUMENTO|S/FRA\.?)\b", "", t)

    # 2️⃣ Mantener solo letras, números y separadores simples
    t = re.sub(r"[^A-Z0-9\-/]", "", t)

    # 3️⃣ Simplificar secuencias repetidas de separadores
    t = re.sub(r"[-/]{2,}", "-", t)

    # 4️⃣ Eliminar separadores iniciales o finales
    t = t.strip("-/")

    # 5️⃣ Casos comunes: “MA-1391”, “MA 1391”, “24-25/MA//1391” → MA1391, 1391, 2425MA1391
    m = re.search(r"([A-Z]{1,5})[-/]?(\d{2,6})", t)
    if m:
        return f"{m.group(1)}{m.group(2)}"

    # 6️⃣ Si son solo dígitos largos
    if re.fullmatch(r"\d{6,}", t):
        return t

    # 7️⃣ Si nada cuadra, devuélvelo limpio
    return t

def _pick_invoice_columns(df: pd.DataFrame) -> list[str]:
    """
    Detecta columnas que probablemente contengan números de factura:
    Ej: 'S/Fra. Número', 'Factura nº', 'Invoice', etc.
    """
    import re
    candidates = [
        "factura", "fra", "invoice", "nº", "numero", "número", "ref", "referenc", "doc"
    ]
    cols = []

    # 1️⃣ Detección por nombre de columna (regex flexible)
    pattern = re.compile("|".join(candidates), re.IGNORECASE)
    for c in df.columns:
        if pattern.search(str(c)):
            cols.append(c)

    # 2️⃣ Si no encuentra por nombre, buscar por patrón en los valores
    if not cols:
        pattern_val = re.compile(r"\b\d{4,}([\/\-\s]?\d{3,})?\b")  # 3020014885 o 2025/137090
        for c in df.columns:
            series = df[c].astype(str).fillna("").head(60)
            hits = sum(bool(pattern_val.search(v)) for v in series)
            if hits / max(1, len(series)) > 0.25:
                cols.append(c)

    return cols[:3]

def _pick_amount_columns(df: pd.DataFrame) -> list[str]:
    cand = ["importe", "total", "bruto", "base", "a pagar", "amount", "total factura", "total euros"]
    cols = []
    for c in df.columns:
        name = str(c).lower()
        if any(k in name for k in cand):
            cols.append(c)
    return cols[:4]

def _first_amount_in_row(row: pd.Series, cols: list[str]) -> float | None:
    for c in cols:
        v = row.get(c, None)
        n = _to_float_eu(v) if v is not None else None
        if n is not None:
            return n
    return None


@app.post("/api/contraste-facturas")
async def contraste_facturas(
    pendientes: UploadFile = File(...),
    facturas: List[UploadFile] = File(...),
):
    import re

    # 1️⃣ Leer Excel
    try:
        pend_df = _read_tabular(pendientes)
        pend_df = _norm_colnames(pend_df)
    except Exception as e:
        return Response(
            content=f"Error leyendo pendientes: {e}",
            media_type="text/plain",
            status_code=400,
        )

    # 2️⃣ Detectar columnas relevantes
    inv_cols = _pick_invoice_columns(pend_df)
    amt_cols = _pick_amount_columns(pend_df)

    if not inv_cols:
        return Response(
            content="No se detectaron columnas de Nº de factura en el Excel de pendientes.",
            media_type="text/plain",
            status_code=400,
        )

    # 3️⃣ Construir índice de facturas en Excel (busca en TODAS las columnas)
    excel_index: dict[str, dict] = {}
    for idx, row in pend_df.iterrows():
        for col in pend_df.columns:
            val = str(row.get(col, "") or "")
            if not val.strip():
                continue
            val_norm = _norm_invoice_code(val)
            if not val_norm:
                continue
            excel_index[val_norm] = {
                "fila": idx + 2,
                "columna": col,
                "valor": val,
            }

    # 4️⃣ Procesar los PDFs SOLO por nombre de archivo
    resultados = []
    for f in facturas:
        if not f.filename.lower().endswith(".pdf"):
            return Response(
                content=f"'{f.filename}' no es un PDF",
                media_type="text/plain",
                status_code=400,
            )
        nombre = f.filename.upper()

           # 1️⃣ Extrae códigos principales normalizados
        matches = re.findall(
            r"(?:INV\d{3,}|\b\d{6,}\b|\d{4}[-_/]\d{3,}|[A-Z]{1,5}[\s\-_/]*\d{2,6}|\d{2,6}[\s\-_/]*[A-Z]{1,5})",
            nombre.upper()
        )

        # Normaliza y separa letras/números pegados (ej: 84927TRAVI → 84927, TRAVI)
        posibles_codigos = []
        for m in matches:
            nm = re.sub(r"[\s\-_/.]", "", m.strip().upper())
            # Separar letras y números adyacentes
            nm_split = re.split(r"(?<=\d)(?=[A-Z])|(?<=[A-Z])(?=\d)", nm)
            for part in nm_split:
                part = part.strip()
                if part and part not in posibles_codigos:
                    posibles_codigos.append(part)

        coincidencia = None
        razon = "No se detectó ningún código en el nombre del archivo"

        # 2️⃣ Primera pasada: comparación normal con Excel
        for code in posibles_codigos:
            code_norm = _norm_invoice_code(code)
            if code_norm in excel_index:
                coincidencia = excel_index[code_norm]
                razon = (
                    f"'{code}' del archivo coincide con "
                    f"celda (columna '{coincidencia['columna']}', fila {coincidencia['fila']}) → {coincidencia['valor']}"
                )
                break

        # 3️⃣ Segunda pasada: solo si no hubo coincidencia, busca cualquier bloque de ≥4 alfanuméricos seguidos
        if not coincidencia:
            extra_blocks = re.findall(r"[A-Z0-9]{4,}", nombre.upper())
            for eb in extra_blocks:
                eb_norm = _norm_invoice_code(eb)
                if eb_norm in excel_index and all(
                    eb_norm != _norm_invoice_code(c["code"]) for c in coincidencias_encontradas
                ):
                    coincidencias_encontradas.append({
                        "code": eb,
                        "coincidencia": excel_index[eb_norm],
                        "origen": "secundaria",
                    })

        if not coincidencia:
            razon = f"Ningún código del archivo ({', '.join(posibles_codigos)}) se encontró en el Excel"

        # ➤ Filtrar años (2020–2039)
        matches = [m for m in matches if not re.fullmatch(r"20[2-3]\d", m)]

        # ➤ Limpiar duplicados
        matches = list(dict.fromkeys(matches))


        # 3️⃣ Normaliza y elimina duplicados
        posibles_codigos = []
        for m in matches:
            nm = _norm_invoice_code(m)
            if nm and nm not in posibles_codigos:
                posibles_codigos.append(nm)

        # 4️⃣ Buscar coincidencias con el Excel (dos pasadas)
        coincidencias_encontradas = []
        razon = "No se detectó ningún código en el nombre del archivo"

        # ➤ Primera pasada: códigos principales normalizados
        for code in posibles_codigos:
            code_norm = _norm_invoice_code(code)
            if code_norm in excel_index:
                coincidencias_encontradas.append({
                    "code": code,
                    "coincidencia": excel_index[code_norm],
                    "origen": "primaria",
                })

        # ➤ Segunda pasada: bloques genéricos (solo si no estaban ya)
        extra_blocks = re.findall(r"[A-Z0-9]{4,}", nombre.upper())
        for eb in extra_blocks:
            eb_norm = _norm_invoice_code(eb)
            if eb_norm in excel_index and all(
                eb_norm != _norm_invoice_code(c["code"]) for c in coincidencias_encontradas
            ):
                coincidencias_encontradas.append({
                    "code": eb,
                    "coincidencia": excel_index[eb_norm],
                    "origen": "secundaria",
                })

        # 🟨 3️⃣ Tercera vía: rescate numérico puro — solo si NO hay coincidencias hasta ahora
        if not coincidencias_encontradas:
            numeric_blocks = re.findall(r"\d{4,}", nombre)
            for nb in numeric_blocks:
                # Ignorar años comunes
                if nb in {"2020","2021","2022","2023","2024","2025","2026"}:
                    continue
                if nb in excel_index:
                    coincidencias_encontradas.append({
                        "code": nb,
                        "coincidencia": excel_index[nb],
                        "origen": "terciaria",
                    })

        # 🟨 3️⃣ Tercera vía: rescate — busca bloques aislados por separadores
        # 🟨 3️⃣ Tercera vía: rescate mejorado — separa letras y números contiguos
        base_clean = re.sub(r"[^A-Z0-9]", " ", nombre.upper())

        # Separa letras y números adyacentes (para evitar "84927TRAVI")
        base_clean = re.sub(r"(?<=\d)(?=[A-Z])", " ", base_clean)
        base_clean = re.sub(r"(?<=[A-Z])(?=\d)", " ", base_clean)

        rescue_blocks = re.findall(r"\b[A-Z0-9]{4,}\b", base_clean)

        for rb in rescue_blocks:
            if rb in {"2020", "2021", "2022", "2023", "2024", "2025", "2026"}:
                continue
            rb_norm = _norm_invoice_code(rb)
            if rb_norm in excel_index and all(
                rb_norm != _norm_invoice_code(c["code"]) for c in coincidencias_encontradas
            ):
                coincidencias_encontradas.append({
                    "code": rb,
                    "coincidencia": excel_index[rb_norm],
                    "origen": "terciaria",
                })

        # ➤ Evaluar resultado combinado
        if coincidencias_encontradas:
            coincidencia = coincidencias_encontradas[0]["coincidencia"]  # toma la primera para mostrar
            encontrados = [
                f"{c['code']} → fila {c['coincidencia']['fila']} ({c['origen']})"
                for c in coincidencias_encontradas
            ]
            razon = " / ".join(
                [f"'{c['code']}' coincide con celda (columna '{c['coincidencia']['columna']}', fila {c['coincidencia']['fila']}) → {c['coincidencia']['valor']}'" for c in coincidencias_encontradas]
            )
        else:
            coincidencia = None
            razon = f"Ningún código del archivo ({', '.join(posibles_codigos)}) se encontró en el Excel"

            # 🟨 3️⃣ Tercera vía: búsqueda separada SOLO para los no coincidentes
            # Genera nueva lista de matches numéricos (solo dígitos, 4 o más)
            matches_tercera = re.findall(r"\d{4,}", nombre)

            for mt in matches_tercera:
                # Ignorar años típicos
                if mt in {"2020","2021","2022","2023","2024","2025","2026"}:
                    continue

                # Coincidencia directa con el Excel (sin normalizar)
                if mt in excel_index:
                    coincidencia = excel_index[mt]
                    razon = (
                        f"'{mt}' (vía numérica pura) coincide con celda "
                        f"(columna '{coincidencia['columna']}', fila {coincidencia['fila']}) → {coincidencia['valor']}"
                    )
                    break

            # Si sigue sin coincidencia, actualizar razón final
            if not coincidencia:
                razon = (
                    f"Ningún código del archivo ({', '.join(posibles_codigos + matches_tercera)}) "
                    f"se encontró en el Excel"
                )


            # 🟨 3️⃣ Tercera vía: nueva búsqueda exclusiva para los no coincidentes
            matches_tercera = re.findall(r"\d{4,}", nombre)
            for mt in matches_tercera:
                # Ignorar años típicos
                if mt in {"2020","2021","2022","2023","2024","2025","2026"}:
                    continue
                # Coincidencia directa o contenida dentro de una clave normalizada
                for key, info in excel_index.items():
                    if mt == key or mt in key:
                        coincidencia = info
                        razon = (
                            f"'{mt}' (vía numérica pura) coincide con celda "
                            f"(columna '{info['columna']}', fila {info['fila']}) → {info['valor']}"
                        )
                        break



        resultados.append({
            "Archivo": f.filename,
            "CodigosDetectados": posibles_codigos,
            "Coincidencia": bool(coincidencia),
            "Razon": razon,
        })

    # 5️⃣ Preparar preview
    preview = {
        "Resumen": {
            "PDFsProcesados": len(resultados),
            "Coincidencias": len([r for r in resultados if r["Coincidencia"]]),
            "Faltantes": len([r for r in resultados if not r["Coincidencia"]]),
        },
        "Coincidencias": [
            {
                "Documento": r["Archivo"],
                "CoincidenciaDetectada": (
                    f"✅ {r['Razon'].split(')')[0] + ')'}"
                    if r["Coincidencia"]
                    else "—"
                ),
            }
            for r in resultados
            if r["Coincidencia"]
        ],
        "Faltantes": [
            f"⚠️ {r['Archivo']} → {r['Razon']}"
            for r in resultados
            if not r["Coincidencia"]
        ],
    }

    return Response(
        content=json.dumps(preview, ensure_ascii=False, indent=2),
        media_type="application/json",
        headers={"X-Preview": json.dumps(preview, ensure_ascii=True)},
    )



@app.post("/api/bankflowpro")
async def bankflowpro(
    extracto: UploadFile = File(...),
    detalle_remesas: UploadFile | None = File(None),
):
    # 1) Leer el extracto (CSV/XLSX)
    try:
        ext_df = _read_tabular(extracto)
        ext_df = _norm_colnames(ext_df)
    except Exception as e:
        return Response(
            content=f"Error leyendo el extracto: {e}",
            media_type="text/plain",
            status_code=400,
        )

    # 2) Detectar columnas mínimas (por contenido)
    col_fecha = _find_col(ext_df, ["fecha operacion", "fecha de operacion", "fecha", "fecha valor"])
    col_concepto = _find_col(ext_df, [
        "concepto", "descripcion", "descripción", "detalle", "concepto ampliado",
        "detalle del movimiento", "observaciones"
    ])

    # Importe puede venir de formas distintas:
    col_importe = _find_col(ext_df, ["importe", "importe eur", "importe operacion", "amount", "importe operación"])

    # Alternativas por doble columna:
    col_cargo = _find_col(ext_df, ["cargo", "debe", "debito", "débito", "debit"])
    col_abono = _find_col(ext_df, ["abono", "haber", "credito", "crédito", "credit"])

    # Columna de signo / tipo
    col_signo = _find_col(ext_df, ["signo", "d/c", "tipo movimiento", "tipo mov", "movimiento"])

    tengo_importe = bool(col_importe or (col_cargo or col_abono))
    if not (col_fecha and col_concepto and tengo_importe):

        return Response(
            content="No se detectaron columnas mínimas (Fecha/Concepto/Importe) en el extracto.",
            media_type="text/plain",
            status_code=400,
        )

    # 3) Construir salida base (Volvemos a usar "Total")
    out_rows = []
    
    for _, r in ext_df.iterrows():
        fecha = _to_date_ddmmyyyy(r.get(col_fecha, "")) if col_fecha else ""
        concepto = str(r.get(col_concepto, "") or "") if col_concepto else ""

        # --- Importe robusto ---
        imp = 0.0
        if col_cargo or col_abono:
            cargo = _to_float_eu(r.get(col_cargo, "")) if col_cargo else 0.0
            abono = _to_float_eu(r.get(col_abono, "")) if col_abono else 0.0
            imp = (abono or 0.0) - (cargo or 0.0)
        elif col_importe:
            imp = _to_float_eu(r.get(col_importe, "")) or 0.0
            if col_signo:
                s = str(r.get(col_signo, "") or "").strip().lower()
                neg = s in {"d", "debe", "cargo", "-", "debito", "débito"}
                pos = s in {"h", "haber", "abono", "+", "credito", "crédito"}
                if neg and imp > 0:
                    imp = -imp
                if pos and imp < 0:
                    imp = -imp
        
        out_rows.append({
            "Fecha": fecha,
            "Concepto": concepto,
            "Tipo": "",
            "Importe": imp,
            "Comisión": 0.0,
            "IVA": 0.0,
            "IRPF": 0.0,
            "Importe Neto": imp, # <-- Nueva columna final es Importe Neto

        })

    # --- Leer detalle remesas (si existe) ---
    rem_df = None
    avisos = []
    if detalle_remesas:
        try:
            rem_df = _read_tabular(detalle_remesas)
            rem_df = _norm_colnames(rem_df)
        except Exception as e:
            avisos.append(f"Aviso: No se pudo leer el detalle de remesas: {e}")

    # --- DataFrame con "Total" ---
    out_df = pd.DataFrame(out_rows, columns=["Fecha", "Concepto", "Tipo", "Importe", "Comisión", "IVA", "IRPF", "Importe Neto"])


    # 3.1) Aplicar pipeline completo (Reglas + Remesas)
    out_df, avisos_bankflow = process_bankflow(out_df, rem_df)
    avisos.extend(avisos_bankflow)

        # 4) X-Preview (máx. 50 filas)
    preview_rows = []

    for _, r in out_df.head(50).iterrows():
        concepto_preview = str(r.get("Concepto", "") or "")
        if len(concepto_preview) > 30:
            concepto_preview = concepto_preview[:30] + "…"

        tipo_str = str(r.get("Tipo", "") or "")

        # parseo robusto (EU) para números en preview
        def _to_num(v):
            if isinstance(v, (int, float)):
                return float(v)
            return _to_float_eu(v) or 0.0

        imp      = _to_num(r.get("Importe", 0.0))
        com_val  = _to_num(r.get("Comisión", r.get("Comision", 0)))  # ← tilde y fallback
        iva_val  = _to_num(r.get("IVA", 0.0))
        irpf_val = _to_num(r.get("IRPF", 0.0))
        neto_val = _to_num(r.get("Importe Neto", 0.0))

        # Caso especial: Comisión bancaria → comisión = importe, sin IVA/IRPF
        if "comisión bancaria" in tipo_str.lower() or "comision bancaria" in tipo_str.lower():
            com_val  = imp
            iva_val  = 0.0
            irpf_val = 0.0
            neto_val = abs(imp)

        preview_rows.append({
            "Fecha": r.get("Fecha", ""),
            "Concepto": concepto_preview,
            "Tipo": tipo_str,
            "Importe": _fmt_eur(imp),
            "Comisión": _fmt_eu(com_val),
            "Comision": _fmt_eu(com_val),

            "IVA": _fmt_eur(iva_val),
            "IRPF": _fmt_eur(irpf_val),
            "Importe Neto": _fmt_eur(neto_val),
        })

    x_preview = json.dumps({"Filas": int(len(out_df)), "Muestra": preview_rows}, ensure_ascii=True)

    # 5) Generar Excel (hoja única Movimientos_desglosados)
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        sheet = "Movimientos_desglosados"
        out_df.to_excel(w, index=False, sheet_name=sheet)

        from openpyxl.styles import Font, Alignment, PatternFill, Border
        from openpyxl.utils import get_column_letter

        wb = w.book
        ws = wb[sheet]

        ws.sheet_view.showGridLines = False

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1f3564")
        left_align = Alignment(horizontal="left", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_align
            cell.border = Border()

        euro_fmt = '#,##0.00'
        # D..H (Importe, Comisión, IVA, IRPF, Importe Neto)
        for col_idx in range(4, 9):

            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter][1:]:
                cell.number_format = euro_fmt
                cell.alignment = left_align

        for col_idx in range(1, 9):  # A..H
            col_letter = get_column_letter(col_idx)
            max_len = 10
            for cell in ws[col_letter]:
                v = cell.value
                if v is None:
                    l = 0
                elif isinstance(v, (int, float)):
                    l = len(f"{v:,.2f}")
                else:
                    l = len(str(v))
                if l > max_len:
                    max_len = l
                if cell.row != 1:
                    cell.border = Border()

            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 50))

        # --- INICIO: Sombrear filas de remesa desglosada ---
        
        # Define el color de sombreado (Azul claro sutil)
        remesa_fill = PatternFill("solid", fgColor="EAF2F8") # <-- NUEVO COLOR

        for idx, row in out_df.iterrows():
            try:
                tipo_lower = str(row.get("Tipo", "") or "").lower()
                comision_val = float(row.get("Comisión", 0.0) or 0.0)

                es_traspaso = "traspaso" in tipo_lower
                es_comision_banco = "comision" in tipo_lower
                
                is_remesa_line = (comision_val == 0.0) and not es_traspaso and not es_comision_banco

                if is_remesa_line:
                    ws_row = idx + 2 
                    for cell in ws[ws_row]:
                        cell.fill = remesa_fill
            except Exception:
                pass
        # --- FIN: Sombreado ---

    xlsx_bytes = out.getvalue()

    headers = {
        "Content-Disposition": 'attachment; filename*=UTF-8\'\'Movimientos_desglosados.xlsx',
        "X-Preview": x_preview,
    }
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


if __name__ == "__main__":
    import uvicorn  # puedes quitar esta línea si ya lo importas arriba
    print("✅ FastAPI corriendo en http://127.0.0.1:8000")
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
